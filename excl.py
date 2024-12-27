import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

class SpreadsheetApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Advanced Spreadsheet")
        self.root.geometry("800x600")

        # Variables to hold the spreadsheet data
        self.data = [[None] * 10 for _ in range(10)]  # 10x10 grid

        # Menu bar
        self.create_menu()

        # Table (Canvas to simulate spreadsheet)
        self.create_table()

    def create_menu(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Open", command=self.load_spreadsheet)
        file_menu.add_command(label="Save", command=self.save_spreadsheet)
        file_menu.add_command(label="Exit", command=self.root.quit)

        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        edit_menu.add_command(label="Bold", command=self.make_bold)
        edit_menu.add_command(label="Sum", command=self.calculate_sum)

    def create_table(self):
        self.table_frame = tk.Frame(self.root)
        self.table_frame.pack(fill=tk.BOTH, expand=True)

        self.cells = []
        for i in range(10):
            row_cells = []
            for j in range(10):
                cell = tk.Entry(self.table_frame, width=10, font=('Arial', 10))
                cell.grid(row=i, column=j, padx=5, pady=5)
                row_cells.append(cell)
            self.cells.append(row_cells)

    def load_spreadsheet(self):
        file_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            try:
                wb = load_workbook(file_path)
                sheet = wb.active
                self.data = [[sheet.cell(row=i+1, column=j+1).value for j in range(10)] for i in range(10)]
                self.update_table()
            except Exception as e:
                messagebox.showerror("Error", f"Error loading file: {e}")

    def save_spreadsheet(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            try:
                wb = Workbook()
                sheet = wb.active
                for i in range(10):
                    for j in range(10):
                        sheet.cell(row=i+1, column=j+1, value=self.data[i][j])
                wb.save(file_path)
            except Exception as e:
                messagebox.showerror("Error", f"Error saving file: {e}")

    def update_table(self):
        for i in range(10):
            for j in range(10):
                self.cells[i][j].delete(0, tk.END)
                self.cells[i][j].insert(tk.END, str(self.data[i][j]) if self.data[i][j] is not None else '')

    def make_bold(self):
        for i in range(10):
            for j in range(10):
                if self.cells[i][j].focus_get():
                    current_font = self.cells[i][j].cget("font")
                    new_font = ('Arial', 10, 'bold' if 'bold' not in current_font else 'normal')
                    self.cells[i][j].config(font=new_font)
                    self.data[i][j] = self.cells[i][j].get()

    def calculate_sum(self):
        for j in range(10):
            try:
                col_sum = 0
                for i in range(10):
                    cell_value = self.data[i][j]
                    if isinstance(cell_value, (int, float)):
                        col_sum += cell_value
                self.cells[9][j].delete(0, tk.END)
                self.cells[9][j].insert(tk.END, str(col_sum))
                self.data[9][j] = col_sum  # Store sum in the last row
            except Exception as e:
                messagebox.showerror("Error", f"Error in calculating sum: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = SpreadsheetApp(root)
    root.mainloop()
